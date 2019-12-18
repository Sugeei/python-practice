# -*- coding: utf-8 -*-


import urllib
import requests
requests = requests.Session()
from lxml import etree
from pymongo import MongoClient
client = MongoClient()
db = client['linpin']
collection = db['linpin']
from openpyxl import Workbook, load_workbook
import logging
logging.basicConfig(format='%(asctime)s: %(levelname)s: %(message)s', filename='./Crawl_log.log', level=logging.INFO)


opportunity_keywords = set((u'数据', u'分析', u'挖掘'))
base_url = 'https://www.liepin.com/zhaopin/?industries=&dqs=&salary=&jobKind=2&pubTime=7&compkind=&compscale=&industryType=' \
      '&searchType=1&clean_condition=&isAnalysis=&init=-1&sortFlag=15&fromSearchBtn=2&head&ckid=c427f8bba5c05707' \
      '&&ckid=c427f8bba5c05707&headckid=c4ed7f85ba1ed471&flushckid=1&key='

campus_base_url = 'https://campus.liepin.com/xycompany/?keywords='
headers = {
    'User-Agent': 'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/48.0.2564.116 '
                  'Safari/537.36',
    'DNT': '1',
    'Host': 'campus.liepin.com',
    'Pragma': 'no-cache',
    'Upgrade-Insecure-Requests': '1'
}

cookies = {
    'Cookie': '__uuid=1476718999772.16; _uuid=D34FB9D1FFA8420C15157A9C17482E83; gr_user_id=53427bcd-f337-4a67-abd8-6cea3885fb65; _fecdn_=1; verifycode=6fa6e5af690a4c3a882e6be4adb2a02d; JSESSIONID=B5C4C3E0D35B3F2F72C7A3575835765E; __tlog=1477208983609.41%7C00000000%7CR000000075%7Cs_o_009%7Cs_o_009; __session_seq=39; __uv_seq=39; _mscid=00000000; Hm_lvt_a2647413544f5a04f00da7eee0d5e200=1476719001,1477208985,1477212163; Hm_lpvt_a2647413544f5a04f00da7eee0d5e200=1477212761'
}


def search_company(company):
    url = campus_base_url + company
    try:
        search_result = requests.get(url, headers=headers, cookies=cookies)
    except:
        logging.info(u'Failed to get Content: %s' % url)
        return None, None
    if search_result and '没有找到符合条件的公司' in search_result.content:
        return None, None
    content = etree.HTML(search_result.content)
    result = content.xpath('//div[@class="search-results-list clearfix"]/dl/dd/p/a')[0].get('href')
    return result,  search_result.cookies


def opportunity_list(url):
    opportunity_result = requests.get(url, headers=headers).content
    content = etree.HTML(opportunity_result)
    results = content.xpath('//div')
    for result in results:
        print result


def show_detail(search_url, url, cookies, company):
    headers['Accept'] = 'application/json, text/javascript, */*; q=0.01'
    headers['Referer'] = search_url
    headers['X-Requested-With'] = 'XMLHttpRequest'
    try:
        detail_results = requests.get(url, headers=headers, cookies=cookies).json().get('data').get('list')
    except:
        return
    for result in detail_results:
        job_info = {}
        if any(kw in result.get('xyjob_title') for kw in opportunity_keywords):
            job_info['xyjob_title'] = result.get('xyjob_title')
            job_info['position'] = result.get('xyjob_dq_name')
            job_info['job_type'] = result.get('xyjob_jobtitle_name')
            job_info['org_name'] = result.get('org_name')
            job_info['count'] = result.get('xyjob_count')
            job_info['xyjob_desc'] = result.get('xyjob_desc')
            job_info['company_name'] = cs
            collection.insert_one(job_info)


def read_company(path):
    wb = load_workbook(path)
    sheet_ranges = wb[u'公司列表最新']
    alllist = []
    for i in xrange(2, 113):
        logging.info(u'Reading %d company' % i)
        alllist.append(sheet_ranges.cell(row=i, column=5).value)
    return alllist


if __name__ == '__main__':
    company_set = read_company('./company.xlsx')
    for cs in company_set:
        search_url, search_cookies = search_company(cs)
        if search_url:
            cid = search_url.split('/')[-2]
            detail_url = 'https://campus.liepin.com/xycompany/getJobList.json?orgId=&dq=&xycompId=' + cid + '&curPage='
            for i in xrange(2):
                show_detail(search_url, detail_url + str(i), cookies=search_cookies, company=cs)
                logging.info('Save %s successfully' % cs)
        else:
            logging.info(u'%s No opportunities for Campus' % cs)
